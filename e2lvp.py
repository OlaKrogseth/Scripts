import openpyxl  # Package for reading excel files (.xlsx) into Python
from itertools import compress
import re  # For reading and processing text strings
import os  # Use for compiling PDF document


def _is_number(s):
    """
    This function tells us if the object "s" is a number or not.
    This is useful for working out if the excel cell contains a number, and hence whether we need to check how many d.p.
    to round to. It checks by trying to convert the string to a float. If it fails, "s" is not a number
    Args:
        s: [string]
    Returns:
        True/False: Answers if "s" is a string containing only a number
    """

    try:
        float(s)
        return True
    except ValueError:
        return False


def _cell_is_value(s):
    """
    We only want to round numbers in the table if it is comes from a cell that is only a number, and not from any
    potential table headers cells than contain a number. So check to see if the cell contains only numbers, decimal
    points, parenthesis, asterisk. In which case, return yes. These other terms are included as often tables report
    standard errors in parenthesis or include asterisk to denote statistical significance.
    I added "e" and "E" into the list of characters to allow for scientific notation. Currently the code rounds off
    scientific notation to the specified number of DPs. Should probably create a setting to allow for scientific
    notation to be preserved.
    :param s: string of a number that is to be rounded
    :return: True/False: Answers if "s" is from a cell primarily a number
    """

    search_fun = re.compile(r'[^0-9*.()+-eE]').search
    return not bool(search_fun(s))


def _clean_cell_str(s):
    """
    Sometimes (especially if the excel spreadsheet is generated by another program), cells contents are actually
    formatted as [="CONTENTS"]. If this is the case, we want to return only the CONTENTS and not the =" at the start
    or " at the end.
    Args:
        s: [string]
    Returns:
        string containing only the true content of the cell
    """

    if re.search('=".*"', s):
        return s[2:-1]
    else:
        return s


def _tupple2latexstring(row_tup, usr_settings, merge_list):
    """
    This function converts a tupple of openpyxl CELLs into a single row string
    of LaTeX code for inclusion in the table. It loops over each cell and appends
    the appropriate text (representing the LaTeX code) to the string which it
    returns at the end.
    :param row_tup: [tupple] contains the openpyxl CELLs for a single row of the table
    :param usr_settings: [dict] user defined options
    :param merge_list: list of cells in the row that are merged together
    :return: A string of the row cells formatted in the LaTeX style.
    """

    num_elements = len(row_tup)  # how many columns we have in the row

    str_out = ""  # initilise the output string

    merge_start_cols = merge_list[0]
    merge_end_cols = merge_list[1]
    merge_match_det = merge_list[2]

    colidx = 0
    multiidx = 0

    while colidx < num_elements:  # for each column/cell in the tupple for the row

        # Check to see if the column/cell is part of a multicolumn/row

        if colidx in merge_start_cols:
            # Multicolumn/row

            multcol_indx = merge_start_cols.index(colidx)
            value_string = _clean_cell_str(merge_match_det[multcol_indx])

            colidx = merge_end_cols[multcol_indx]

            multiidx += 1

        else:
            # Get the main text for that cell.

            #########
            # Step 1: Get the "value_string" giving the text displayed in the cell
            #########

            if row_tup[colidx].value is None:
                # In this case, the cell is empty, so

                value_string = " "  # Cell is empty of value

            else:  # Case when the cell contains something
                # Get content of cell, and if needed, apply the d.p. rounding rule to the content.
                if usr_settings['roundtodp']:

                    if _cell_is_value(str(row_tup[colidx].value)):
                        value_string = _round_num_in_str(_clean_cell_str(str(row_tup[colidx].value)),
                                                         usr_settings['numdp'])
                    else:
                        value_string = _clean_cell_str(str(row_tup[colidx].value))
                else:
                    value_string = _clean_cell_str(str(row_tup[colidx].value))

            #########
            # Step 2: Apply formatting to the cell's value
            #########

            # The cell might have special formatting applied to the value inside it (e.g. bold text).
            # Apply the LaTeX version of this formatting to the string

            # Apply bold font if needed
            if row_tup[colidx].font.__dict__['b']:
                value_string = "\\textbf{" + value_string + "}"

            # Apply italicize if needed
            if row_tup[colidx].font.__dict__['i']:
                value_string = "\\textit{" + value_string + "}"

            # Apply font color
            if isinstance(row_tup[colidx].font.color.rgb, str):
                value_string = "\\textcolor[HTML]{" + row_tup[colidx].font.color.rgb[2:] + "}{" + value_string + "}"

            # Cell background color
            if isinstance(row_tup[colidx].fill.start_color.index, int):  # built in color

                # Currently cannot handle this case
                value_string = value_string

            else:

                if row_tup[colidx].fill.start_color.index is not '00000000':
                    value_string = "\\cellcolor[HTML]{" + row_tup[colidx].fill.start_color.index[2:] + "}{" + \
                                   value_string + "}"

        #########
        # Step 3: Now that we have to LaTeX code for that cell/column, append it to the string for the entire row.
        #########

        # Append formatted string for this cell to the string out
        str_out += value_string

        # If this isnt the last element, add cell divider
        if colidx < num_elements - 1:
            str_out += " \t & \t "

        colidx += 1

    # Now that we have looped over all elements, add on line ending code for the end of the row string
    str_out += " \\\ \n"

    return str_out


def _check_for_vline(col_tup, loc):
    """
    Look for vertical lines down the entire length of the column.
    We do this by looping over all the cells, and then counting how many of them have a vertical line in location "loc"
    :param col_tup:  [tupple] contains the openpyxl CELLs for a single row of the table
    :param loc: [string] 'top' or 'bottom'
    :return:
    """

    num_rows = len(col_tup)  # Number of rows in the column

    count = 0  # Initialise count

    for rownum in range(0, num_rows):  # For each row

        # Check to see if there is a border style in location "loc"
        if col_tup[rownum].border.__dict__[loc].border_style is not None:
            # Add one to our count
            count += 1

    # Check to see if every row has a border style in location "loc"
    if count == num_rows:
        return True
    else:
        return False


def _create_cline_code(cell_has_rule_bool, booktabs=False):
    """
    Creates the code for horizontal lines that do not span the entire length of the table, only a few cells.
    :param cell_has_rule_bool: [list] whose elements are True/False for each cell indicating whether the horizontal rule
    includes this cell.
    :param booktabs:  True/False. Should the code return code for the booktabs package or regular LaTeX?
    :return: A string containing the code needed to draw the horizontal lines.
    E.g. "\cmidrule(r){1-4} \cmidrule(r){6-9} \n"
    """

    # Initialize the output string
    str_out = ''

    num_column = len(cell_has_rule_bool)  # How many elements in the row

    # Create a flag to indicate whether we are starting a new cmidrule/cline or not.
    look_for_new_crule = True  # we are looking for the next True to start a new \cmidrule or \cline

    for colind in range(0, num_column):  # For each column/cell

        if (cell_has_rule_bool[colind] is True) & (look_for_new_crule is True):
            # The current cell has a cmidrule/cline, and we are looking for the next crule. So start a new rule/line in
            # the LaTeX code

            colnum = colind + 1  # column number is one more than the python index

            # Append new line/rule to str_out
            if booktabs is True:
                str_out += '\\cmidrule(r){' + str(colnum) + '-'
            else:
                str_out += '\\cline{' + str(colnum) + '-'

            # Turn off flag since now we are going to be looking for where this particular cline ends
            look_for_new_crule = False

            continue  # Move on to next table cell

        elif (cell_has_rule_bool[colind] is False) & (look_for_new_crule is True):
            # This is the case when we are searching for a new cline to begin, but the current table cell does not
            # contain a crule, so we carry on looking
            continue

        elif cell_has_rule_bool[colind] is False:
            # We have found a column/cell without a cline (=False), and we are looking to close the currently opened
            # cline/crule as look_for_new_crule=False (because the cases where look_for_new_crule=True are dealt with
            # by the previous elif case). Therefore, we want to add the LaTeX code to close the current crule/cline

            str_out += str(colind) + '} \t '  # colidx = colnum-1, which is the last column to include

            look_for_new_crule = True  # Turn flag back on so we are searching for the next crule start
            continue

        elif (cell_has_rule_bool[colind] is True) & (colind == num_column - 1):
            # If we get to the end of the table, and the column/cell still has a cline, end the cline.
            # last one is True

            str_out += str(num_column) + '} \t '

    # The above cases exhaust all possibilities, so no need for "else" statement

    # End the LaTeX line and return the string
    str_out += ' \n'

    return str_out


def _create_horzrule_code(row_tup, loc, merge_start_cols, merge_end_cols, usr_settings):
    """
    Create LaTeX code for horizontal lines, above or below (defined by 'loc'), that particular row_tup.
    Horizontal lines may either span the entire width of the table, or along a few columns.
    Args:
        row_tup: [tuple] a particular row of cells.
        loc: [string] either 'top' or 'bottom' to indicate where (relative to this particular
                row) we should check for any horizontal lines.
        merge_start_cols: [list] the row numbers related to where merged cells start.
        merge_end_cols: [list] the row numbers related to where merged cells stop.
        usr_settings: [dictionary] user settings - tells us whether to use booktabs code or not.
    Returns:
        A string containing the LaTeX code needed to draw the horizontal line(s) for that particular row.
    """

    num_column = len(row_tup)  # number of columns/elements in this particular row

    # Step 1: Find which cells have horizontal rules
    # Construct a list with True/False elements to indicate if the horizontal rule applies to that cell.

    cell_has_rule = []  # Pre-allocate list

    for colnum in range(0, num_column):  # for each column in the row

        # Check to see if this particular column falls within the span of merged cells
        # If so, we can ignore it, and use the details from the first cell of the merged cells
        cond_1 = [colnum > x for x in merge_start_cols]
        cond_2 = [colnum <= x for x in merge_end_cols]

        cond_combine = []
        for i in range(0, len(cond_1)):
            cond_combine.append(cond_1[i] is True and cond_2[i] is True)

        if any(x is True for x in cond_combine):  # cell is a subsequent merged cell
            cell_has_rule.append(cell_has_rule[-1])

        else:

            if row_tup[colnum].border.__dict__[loc].border_style is not None:
                cell_has_rule.append(True)
            else:
                cell_has_rule.append(False)

    if sum(cell_has_rule) == 0:  # If there are no rules and any cell, there is no line here, so return a blank string

        return ''

    else:  # There exists some horizontal rule on at least part of the row, so return the appropriate LaTeX code

        # If user has specified booktabs
        if usr_settings['booktabs'] is True:

            if sum(cell_has_rule) == num_column:
                return '\midrule \n'
            else:
                return _create_cline_code(cell_has_rule, booktabs=True)

        else:

            if sum(cell_has_rule) == num_column:
                return '\hline \n'
            else:
                return _create_cline_code(cell_has_rule, booktabs=False)


def _get_merged_cells(sheet):
    """
    Locate all the merged cells within a sheet, return the row and column locations of the start and end, and also
    return the LaTeX code for the merged cells.
    :param sheet: [tuple] openpyxl excel worksheet object
    :return: list containing (1) index of the rows of each merged cells first cell, (2) index of the column of each
    merged cell first cell, (3) index of the columns of each moerged cells last cell, (4) index of the column of each
    merged cell last cel, (5) LaTeX code for the
    """

    start_row = []
    start_col = []

    end_row = []
    end_col = []

    latex_code = []

    if len(sheet.merged_cell_ranges) == 0:
        return [[], [], [], [], []]  # No merged cells, so return an empty list

    for merge_ in sheet.merged_cell_ranges:  # For each merge in the sheet

        # Split the location string of the merge, and convert it it to an index number (e.g. "A3")
        merge_loc_str = re.split(':', merge_)

        # convert string to col/row index numbers
        start_coord = openpyxl.utils.coordinate_to_tuple(merge_loc_str[0])
        end_coord = openpyxl.utils.coordinate_to_tuple(merge_loc_str[1])

        start_row.append(start_coord[0] - 1)
        start_col.append(start_coord[1] - 1)

        end_row.append(end_coord[0] - 1)
        end_col.append(end_coord[1] - 1)

        value_string = sheet[merge_loc_str[0]].value

        if sheet[merge_loc_str[0]].font.__dict__['b']:
            value_string = "\\textbf{" + value_string + "}"

        # Apply italicize if needed
        if sheet[merge_loc_str[0]].font.__dict__['i']:
            value_string = "\\textit{" + value_string + "}"

        # Get span of multicolumn
        multi_col_length = end_coord[1] - start_coord[1] + 1

        # Get alignment
        halign = sheet[merge_loc_str[0]].alignment.__dict__['horizontal'][0]  # get the first letter

        latex_code.append('\multicolumn{' + str(multi_col_length) + '}{' + halign + '}{' + value_string + '}')

    return [start_row, start_col, end_row, end_col, latex_code]


def _pick_col_text_alignment(col_tup):
    """
    For a given column, choose the alignment (left, center, right) based
    on the alignment choice of the majority of the cells
    Args:
        col_tup: [tuple] containing a column of the table.
    Returns:
        A string ('l'/'c'/'r') indicating the alignment to use
    """

    max_column = len(col_tup)

    # Preallocate counters
    count_left = 0
    count_center = 0
    count_right = 0

    # Loop over each row, and count the alignment types
    for rn in range(0, max_column):

        # If the user doesnt speicify an alignment in Excel, we see the alignment
        # choice as "None". So let us assign default values. If a number, align
        # right, if not, align left.

        if col_tup[rn].alignment.__dict__['horizontal'] is None:

            # Check to see if the value is a number
            if col_tup[rn].value is None:
                align_val = 'ignore'
            elif _is_number(col_tup[rn].value):
                align_val = 'right'
            else:
                align_val = 'left'

        else:
            align_val = col_tup[rn].alignment.__dict__['horizontal']

        if align_val in ['left']:

            count_left += 1

        elif align_val in ['center']:

            count_center += 1

        elif align_val in ['right']:

            count_right += 1

    # Find the maximum, in the case of a tie, we break the tie by the order: L,C,R
    max_count = max([count_left, count_center, count_right])

    if count_left == max_count:
        return 'l'
    elif count_center == max_count:
        return 'c'
    elif count_right == max_count:
        return 'r'


def _round_num_in_str(str_in, num_dp):
    """
    For a given string, round any number to the appropriate number of d.p.
    Args:
        str_in: [string] string containing numbers to round
        num_dp: [scalar] number of decimal places to round each number to
    Returns:
        A string where the numbers in str_in have been rounded.
    """

    str_out = str_in

    # Extract a list of all numbers in the string
    list_found_num_str = re.findall("\d+[.]\d*", str_in)  # Add a question mark behind the "]" to round all numbers
    # (even those without a DP)

    list_found_scientific_num_str = re.findall("\d+[.]*\d*e[+-]\d*", str_in)

    list_found = list_found_num_str + list_found_scientific_num_str

    # Create a list of the found numbers rounded to the appropriate d.p.
    list_nums = [float(s) for s in list_found]

    str_format = '%.' + str(num_dp) + 'f'

    for ii in range(0, len(list_found)):
        # For each number found, substitute in the rounded number
        str_out = re.sub(list_found[ii], str_format % list_nums[ii], str_out)

    return str_out


def _all_nones(iterable):
    """
    Tells us if every value within the tuple iterable is None (missing)
    :param iterable:
    :return:
    """

    for element in iterable:
        if element.value is not None:
            return False
    return True


def _get_table_dimensions(sheet):
    """
    The table within the sheet may not start in cell A1. This function finds the location of the table within the sheet
    by looking for the upper-left and bottom-right most cells that have content. It returns the location of these two
    corner cells.
    :param sheet: Excel worksheet object
    :return:    start_row_idx: row number of the upper-left most cell that contains something
                start_col_idx: column number of the upper-left most cell that contains something
                end_row_idx: row number of the bottom-right most cell that contains something
                end_col_idx: column number of the bottom-right most cell that contains something
    """

    # Pre-allocate starting indices
    start_col_idx = 0
    start_row_idx = 0

    # Pre-allocate end indices (adjust for python starting index at zero, and Excel starting at 1)
    end_col_idx = sheet.max_column - 1
    end_row_idx = sheet.max_row - 1

    # Trim off any empty columns at the end of the table
    for col_num in range(sheet.max_column - 1, -1, -1):

        if _all_nones(list(sheet.columns)[col_num]):
            # Trim the column for the sheet
            end_col_idx = end_col_idx - 1
        else:
            # current final column has a value so stop trimming
            break

    # Trim off any empty rows at the end of the table
    for row_num in range(sheet.max_row - 1, -1, -1):

        if _all_nones(list(sheet.rows)[row_num]):
            # Trim the column for the sheet
            end_row_idx = end_row_idx - 1
        else:
            # current final column has value so stop trimming
            break

    # Trim off any empty columns at the start of the table
    for col_num in range(0, sheet.max_column):

        if _all_nones(list(sheet.columns)[col_num]):
            # Trim the column for the sheet
            start_col_idx = start_col_idx + 1
        else:
            # current final column has value so stop trimming
            break

    # Trim off any empty rows at the start of the table
    for row_num in range(0, sheet.max_row):

        if _all_nones(list(sheet.rows)[row_num]):
            # Trim the column for the sheet
            start_row_idx = start_row_idx + 1
        else:
            # current final column has value so stop trimming
            break

    return start_row_idx, start_col_idx, end_row_idx, end_col_idx


def _create_column(table_in, col_idx):
    """
    Takes a table and returns a tuple containing only a single column of the table. Useful for analysing a single
    column of the table
    :param table_in: table to extract column from
    :param col_idx: index of the column to be extracted
    :return: tuple of just column col_idx
    """

    nrows = len(table_in)  # number of rows in the table

    col = []  # preallocate

    for row_num in range(0, nrows):

        col += [table_in[row_num][col_idx]]

    return tuple(col)


def create_pdf_of_tables(workbook, output_dir):
    """
    Write and compile a LaTeX document of all the tables contained within the workbook. This is useful way to quickly
    check all the output looks good
    :param workbook: openpyxl workbook object
    :param output_dir: [string] directory of where the output should be stored
    :return: none. Complies PDF in output directory
    """

    tex_file = open(output_dir + '/output_all_tables.tex', 'w')

    # Write LaTeX preamble

    tex_file.write('\\documentclass[12pt]{article}\n\n')
    tex_file.write('\\usepackage{booktabs}\n')
    tex_file.write('\\usepackage[table]{xcolor}\n')
    tex_file.write('\\usepackage{parskip}\n')

    tex_file.write('\n\\begin{document}\n\n')

    # Write Each table to the file
    flag_first_table = True
    for sheet_name in workbook.get_sheet_names():

        if flag_first_table is False:
            tex_file.write('\\newpage\n')
        else:
            flag_first_table = False

        tex_file.write('Table: ' + sheet_name.replace('_', '\_') + '\n\n')
        tex_file.write('\\input{' + sheet_name + '.tex}\n\n')

    # Close file
    tex_file.write('\\end{document}')
    tex_file.close()

    # Compile PDF and put in output directory
    os.chdir(output_dir)
    os.system('pdflatex "' + output_dir + '/output_all_tables.tex"')

    # Clean up temp files
    os.remove(output_dir + '/output_all_tables.aux')
    os.remove(output_dir + '/output_all_tables.log')


def excel2latexviapython(input_excel_filename, output_dir, booktabs=True, includetabular=True, roundtodp=True, numdp=3,
                         makepdf=False):
    """
    This function takes an excel workbook of tables, and creates individual TeX files for the tables found within each
    worksheet of the workbook.
    :param input_excel_filename: [string] path and file name of the excel file containing the tables
    :param output_dir: [string] path of the directory to output the TeX files to
    :param booktabs: [True/False] Should booktabs be used rather than regular horizontal rules?
    :param includetabular: [True/False] Should each table be wrapped in a tabular environment?
    :param roundtodp: [True/False] Should numbers be rounded to a specific number of decimal places?
    :param numdp: [Int] How many decimal places to use (only applies is roundtodp=True)
    :param makepdf: [True/False] Should the code also create a simple PDF document of all the tables?
    :return: None
    """

    # Store the user settings in a dictionary to use
    usr_settings = {'booktabs': booktabs, 'includetabular': includetabular, 'roundtodp': roundtodp, 'numdp': numdp,
                    'makepdf': makepdf}

    # PREAMBLE
    # ==================================================================================================================

    # Print output so user can follow progress.
    print('EXCEL 2 LATEX VIA PYTHON')
    print('Creates .TeX table files from excel file.')
    print('\nSource file:      ' + input_excel_filename)

    # Load in the Excel workbook/file
    workbook = openpyxl.load_workbook(filename=input_excel_filename, data_only=True)

    print('Output directory: ' + output_dir + '\n')
    print('User settings:')
    print('    booktabs: ' + str(usr_settings['booktabs']))
    print('    includetabular: ' + str(usr_settings['includetabular']))
    print('    roundtodp: ' + str(usr_settings['roundtodp']))
    print('    numdp: ' + str(usr_settings['numdp']))
    print('\n')
    print('Starting to create TeX tables (output name, table location within excel sheet')

    # MAIN CODE
    # ==================================================================================================================

    for sheet_name in workbook.get_sheet_names():  # Loop over every worksheet/tab within the input workbook

        # Get the worksheet object for this iteration of the loop
        sheet = workbook[sheet_name]

        # The table within the sheet may not start in cell A1. So find the location of the upper-left and bottom-right
        # corner cells of the table within the sheet
        start_row_idx, start_col_idx, end_row_idx, end_col_idx = _get_table_dimensions(sheet)

        # Get the excel cell labels of the upper-left and bottom-right cells of the table
        start_cell_label = list(sheet.rows)[start_row_idx][start_col_idx].coordinate
        end_cell_label = list(sheet.rows)[end_row_idx][end_col_idx].coordinate

        # Get the number of columns and rows in the table
        num_cols = end_col_idx - start_col_idx + 1
        num_rows = end_row_idx - start_row_idx + 1

        # Trim sheet object down to just the range we care about and store this in a tuple
        table_tuple = tuple(sheet[start_cell_label:end_cell_label])

        # Print to the terminal the name of the table file that is being created this iteration and the excel cells
        # being used to create it
        print('    ' + sheet_name + '.tex    ' + list(sheet.rows)[start_row_idx][start_col_idx].coordinate + ':'
              + list(sheet.rows)[end_row_idx][end_col_idx].coordinate)

        # Create .tex file we will write to
        file = open(output_dir + sheet_name + '.tex', 'w')

        # Preamble of the individual table
        # --------------------------------

        # If the user requested the booktabs options, add a reminder (as a LaTeX comment) to the top of the table that
        # the user will need to load up the package in the preamble of their file.
        if usr_settings['booktabs']:
            file.write('% Note: make sure \\usepackage{booktabs} is included in the preamble \n')

        file.write('% Note: If your table contains colors, make sure \\usepackage[table]{xcolor} is included in the '
                   'preamble \n')

        # If the user wants the table rows wrapped in the tabular environment, write the start of the begin environment
        # command to the output tex file
        if usr_settings['includetabular']:

            col_align_str = "\\begin{tabular}{"  # Preallocate string

            # For each column of the table, append to "col_align_str" any vertical dividers and alignment code for the
            # column
            for colnum in range(0, num_cols):

                # Create column to analyze from the table
                col2a = _create_column(table_tuple, colnum)

                # check to see if there is a vline left of column
                if _check_for_vline(col2a, 'left'):
                    col_align_str += '|'

                # Choose the alignment (l,c,r) of the column based on the majority of alignments in the column's cells
                col_align_str += _pick_col_text_alignment(col2a)

                # check to see if there is a vline right of column
                if _check_for_vline(col2a, 'right'):
                    col_align_str += '|'

            # Create code to write to tex output file
            begin_str = str(col_align_str) + "} \n"

            # Write the \begin{tabular}{*} code to the tex file
            file.write(begin_str)

        # Body of the individual table
        # ----------------------------

        # Find any merged cells within this particular worksheet
        merged_details_list = _get_merged_cells(sheet)

        # Adjust the merged_details_list values for the fact that the table might not start in cell A1
        merged_details_list[0] = [x - start_row_idx for x in merged_details_list[0]]  # start_row
        merged_details_list[1] = [x - start_col_idx for x in merged_details_list[1]]  # start_col
        merged_details_list[2] = [x - start_row_idx for x in merged_details_list[2]]  # end_row
        merged_details_list[3] = [x - start_col_idx for x in merged_details_list[3]]  # end_col

        # For each row in the table's body create a string containing the tex code for that row and write to the output
        # file
        for row_num in range(0, num_rows):

            # Generate list of True/False values to see if they match the row
            elem_picker = [True if item in [row_num] else False for item in merged_details_list[0]]

            # Pick out the column number and mutlicolumn/row details corresponding to this row
            merge_start_cols = list(compress(merged_details_list[1], elem_picker))
            merge_end_cols = list(compress(merged_details_list[3], elem_picker))
            merge_match_det = list(compress(merged_details_list[4], elem_picker))

            # If there is a horizontal rule across all cells at the top, add it to the table
            hrule_str = _create_horzrule_code(table_tuple[row_num], 'top', merge_start_cols, merge_end_cols,
                                              usr_settings)

            # If user requested booktabs, and this is the first row, use toprule rather than midrule
            if (row_num == 0) & usr_settings['booktabs']:
                hrule_str = hrule_str.replace('\\midrule', '\\toprule')

            file.write(hrule_str)

            # Get string of rows contents
            str_2_write = _tupple2latexstring(table_tuple[row_num], usr_settings, [merge_start_cols, merge_end_cols,
                                                                                   merge_match_det])

            # Write row string to file
            file.write(str_2_write)

            # Add any horizontal rule below the row
            hrule_str = _create_horzrule_code(table_tuple[row_num], 'bottom', merge_start_cols, merge_end_cols,
                                              usr_settings)

            # If user requested booktabs, and this is the final row, use bottomrule rather than midrule
            if (row_num == num_rows - 1) & usr_settings['booktabs']:
                hrule_str = hrule_str.replace('\\midrule', '\\bottomrule')

            file.write(hrule_str)

        # Postamble of the individual table
        # ---------------------------------
        if usr_settings['includetabular']:
            # User has requested tabular environment wrapped around the table rows, so end the table
            file.write("\\end{tabular}")

        file.close()  # Close off the current .tex file (completing the creation of the table code)

    # Make PDF of the tables for checking purposes
    if makepdf & includetabular:  # can only compile the tables if the tabular environment is included
        create_pdf_of_tables(workbook, output_dir)

    print('\nCode has completed running')